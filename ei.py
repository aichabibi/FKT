import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px
import plotly.graph_objects as go
import io
import json
from datetime import datetime, timedelta
from openpyxl.styles.colors import COLOR_INDEX
import random
import re
from colorsys import rgb_to_hsv
import base64
import tempfile
import plotly.io as pio
from io import BytesIO
from PIL import Image
import pdfkit

# Carte des couleurs avec leurs descriptions et codes hexadécimaux pour l'affichage
color_map = {
    'FFFFFF': {'desc': 'FKT à venir', 'hex': '#FFFFFF'},
    '92D050': {'desc': 'FKT réalisée', 'hex': '#92D050'},
    '00B0F0': {'desc': 'FKT contrôlée sur Plan', 'hex': '#00B0F0'},
    '92CDDC': {'desc': 'FKT contrôlée sur Terrain', 'hex': '#92CDDC'},
    'FF0000': {'desc': 'FKT avec écart', 'hex': '#FF0000'},
    'FFFF00': {'desc': 'FKT manquante', 'hex': '#FFFF00'},
    'FFC000': {'desc': 'FKT provisoire', 'hex': '#FFC000'},
}


def normalize_hex_color(hex_color):
    """Normalise le code hexadécimal des couleurs pour assurer la cohérence"""
    if not hex_color:
        return None
    if not isinstance(hex_color, str):
        hex_color = str(hex_color)
    hex_color = hex_color.replace('#', '')
    if len(hex_color) == 8:
        hex_color = hex_color[-6:]  # conserve les 6 derniers caractères (ignore les 2 premiers)
    return hex_color.upper().zfill(6)

def safe_extract_hex(color_obj):
    """Extrait un code hexadécimal propre depuis un objet couleur openpyxl"""
    try:
        if hasattr(color_obj, 'rgb') and color_obj.rgb:
            rgb_str = str(color_obj.rgb)
            return rgb_str[-6:].upper()
        elif hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
            from openpyxl.styles.colors import COLOR_INDEX
            if color_obj.indexed < len(COLOR_INDEX):
                return COLOR_INDEX[color_obj.indexed][-6:].upper()
        elif hasattr(color_obj, 'theme') or hasattr(color_obj, 'tint'):
            # fallback pour les couleurs basées sur des thèmes : non gérable de manière fiable ici
            return "FFFFFF"
    except:
        pass
    return "FFFFFF"

def get_color_name(color_obj):
    """Extrait le code hexadécimal de l'objet couleur"""
    if color_obj is None:
        return 'FFFFFF'  # Par défaut, si pas de couleur, on considère blanc
    try:
        hex_val = color_obj.rgb
    except AttributeError:
        return 'FFFFFF'  # Par défaut si pas d'attribut rgb
    
    normalized = normalize_hex_color(hex_val)
    return normalized if normalized else 'FFFFFF'
def find_closest_color(hex_color, color_map, tolerance=10):
    """
    Trouve la couleur la plus proche du hex_color parmi les couleurs du color_map,
    en tolérant une différence sur chaque composante RGB (tolérance = 0 à 255).
    """
    def hex_to_rgb(h):
        h = h.strip("#")
        return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

    rgb_target = hex_to_rgb(hex_color)

    for ref_code in color_map:
        rgb_ref = hex_to_rgb(color_map[ref_code]['hex'].replace("#", ""))
        if all(abs(c1 - c2) <= tolerance for c1, c2 in zip(rgb_target, rgb_ref)):
            return ref_code  # retourne la clé (ex: '92CDDC')

    return 'FFFFFF'  # par défaut : blanc (FKT à venir)

from math import sqrt
def hex_to_rgb(hex_color):
    """Convertit un code hexadécimal (ex: '92CDDC') en tuple RGB (r, g, b)."""
    hex_color = hex_color.strip().replace('#', '').upper()
    if len(hex_color) == 8:
        hex_color = hex_color[-6:]  # ignore l’alpha si présent
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def color_distance(rgb1, rgb2):
    """Distance euclidienne entre deux couleurs RGB."""
    return sum((a - b) ** 2 for a, b in zip(rgb1, rgb2)) ** 0.5

def get_standardized_color(color_code, tolerance=60):
    """
    Associe un code couleur à celle la plus proche dans color_map en tolérant les nuances.
    """
    if not color_code:
        return 'FFFFFF'

    hex_code = color_code.replace('#', '').upper()
    if len(hex_code) == 8:
        hex_code = hex_code[-6:]

    try:
        r1, g1, b1 = [int(hex_code[i:i+2], 16) for i in (0, 2, 4)]
    except:
        return 'FFFFFF'

    def dist(c1, c2):
        return sum((a - b) ** 2 for a, b in zip(c1, c2)) ** 0.5

    min_distance = float('inf')
    closest_color = 'FFFFFF'

    for ref_code in color_map:
        r2, g2, b2 = [int(ref_code[i:i+2], 16) for i in (0, 2, 4)]
        d = dist((r1, g1, b1), (r2, g2, b2))
        if d < min_distance:
            min_distance = d
            closest_color = ref_code

    return closest_color if min_distance <= tolerance else 'FFFFFF'




def is_conforme(value):
    if value is None:
        return None  # Non renseigné

    value_str = str(value).strip().lower()

    non_conforme_keywords = [
        "non conforme", "non-conforme", "Non Conforme", "écart", "ecart",
        "ko", "non", "refusé", "refuse", "incorrect"
    ]
    conforme_keywords = [
        "conforme", "ok", "validé", "valide", "oui", "correct", "accepté", "accepte"
    ]

    # Match partiel au lieu de égalité stricte
    if any(keyword in value_str for keyword in non_conforme_keywords):
        return False
    if any(keyword in value_str for keyword in conforme_keywords):
        return True

    return None


@st.cache_data
def get_sheet_names(file_bytes):
    """Récupère les noms des feuilles du fichier Excel avec mise en cache"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    return wb.sheetnames

def find_header_row(sheet, search_terms=None, max_search_rows=20):
    """Cherche automatiquement la ligne d'en-tête en se basant sur des termes de recherche"""
    if search_terms is None:
        search_terms = ["réglage", "armement", "amortissement"]
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_search_rows, values_only=True), 1):
        found_terms = sum(any(term in str(cell).lower() for term in search_terms) 
                          for cell in row if cell)
        if found_terms >= 2:
            return row_idx
    
    return 1  # Valeur par défaut

@st.cache_data
def load_targeted_data(file_bytes, sheet_name, start_row=None, auto_detect_headers=True):
    """Charge les données à partir du fichier Excel avec mise en cache - version optimisée"""
    # Utiliser read_only=True pour charger plus rapidement
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)  # ⚠️ sans read_only !
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        # Traitement spécial si la feuille est "AIG"
    if sheet_name.lower() == "AIG":
        amortissement_col_index = None
        headers_row = find_header_row(sheet)
        header_cells = next(sheet.iter_rows(min_row=headers_row, max_row=headers_row, values_only=True))

        for idx, cell in enumerate(header_cells):
            if cell:
                cell_clean = str(cell).strip().lower()
                if "Amortissement" in cell_clean:
                    amortissement_col_index = idx
                    break

        if amortissement_col_index is not None:
            conforme_count = 0
            for row in sheet.iter_rows(min_row=headers_row + 1, values_only=True):
                if amortissement_col_index < len(row):
                    cell_value = row[amortissement_col_index]
                    if is_conforme(cell_value):
                        conforme_count += 1

            st.success(f" Feuille AIG : {conforme_count} valeurs 'conforme' trouvées dans la colonne Amortissement.")
        else:
            st.warning(" Colonne 'Amortissement' non trouvée dans la feuille AIG (vérifie les espaces ou retours à la ligne dans l'en-tête).")

    # Traitement spécial si la feuille est "AIG"
        if sheet_name.lower() == "aig":
            amortissement_col_index = None
            headers_row = find_header_row(sheet)

            # Lire la ligne d’en-tête
            header_cells = next(sheet.iter_rows(min_row=headers_row, max_row=headers_row, values_only=True))

            # Afficher les en-têtes nettoyés pour debug
            debug_headers = []
            for idx, cell in enumerate(header_cells):
                if cell:
                    cleaned = re.sub(r'\s+', ' ', str(cell)).strip().lower()
                    debug_headers.append(f"[{idx}] '{cleaned}'")
                    if "amortissement" in cleaned:
                        amortissement_col_index = idx
                else:
                    debug_headers.append(f"[{idx}] None")

            st.info("En-têtes détectés dans la feuille AIG :\n" + "\n".join(debug_headers))

            if amortissement_col_index is not None:
                conforme_count = 0
                for row in sheet.iter_rows(min_row=headers_row + 1, values_only=True):
                    if amortissement_col_index < len(row):
                        cell_value = row[amortissement_col_index]
                        if is_conforme(cell_value):
                            conforme_count += 1

                st.success(f"Feuille AIG : {conforme_count} valeurs 'conforme' trouvées dans la colonne Amortissement.")
            else:
                st.warning("Colonne 'Amortissement' non trouvée dans la feuille AIG malgré nettoyage. Vérifie manuellement les noms ou envoie-moi l’export des headers.")


    # Détection automatique de la ligne d'en-tête si demandé
    if auto_detect_headers and start_row is None:
        start_row = find_header_row(sheet)
    elif start_row is None:
        start_row = 1
    
    # Extraction des en-têtes
    headers = []
    for row in sheet.iter_rows(min_row=start_row, max_row=start_row, values_only=True):
        for col_idx, val in enumerate(row):
            headers.append(str(val).strip() if val else f"Col{col_idx+1}")
        break
    
    # Recherche des colonnes pertinentes
    réglage_cols = []
    amortissement_cols = []
    # Recherche des colonnes de réglage et d'amortissement
    for i, header in enumerate(headers):
        header_lower = header.lower() if isinstance(header, str) else ""
        
        if ("réglage" in header_lower and "armement" in header_lower) or \
           ("reglage" in header_lower and "armement" in header_lower):
            réglage_cols.append((i, header))
        
# Recherche des colonnes d'amortissement
    amortissement_keywords = [
        "amortissement", "amortissement sncf", "conformité", "conformite", "conforme", "type"
    ]

    # Dictionnaire pour tracker toutes les colonnes possibles
    amortissement_candidates = []

    for i, header in enumerate(headers):
        if not isinstance(header, str):
            continue
        cleaned = re.sub(r'\s+', ' ', header.strip().lower())

        if any(keyword in cleaned for keyword in amortissement_keywords):
            amortissement_candidates.append((i, header))

    # Si plusieurs colonnes candidates et certaines sont des copies (.1, .2...) → on garde que l’originale
    if len(amortissement_candidates) > 1:
        amortissement_cols = [
            (i, h) for i, h in amortissement_candidates if not re.search(r'\.\d+$', str(h))
        ]
        # Si on a tout éliminé par erreur, on garde au moins la première
        if not amortissement_cols:
            amortissement_cols = [amortissement_candidates[0]]
    else:
        amortissement_cols = amortissement_candidates



    # Si aucune colonne spécifique n'a été trouvée
    if not réglage_cols:
        for i, header in enumerate(headers):
            if i > 0 and isinstance(header, str) and len(header) > 1:
                réglage_cols.append((i, header))

    # Initialisation des compteurs
    color_summary = {header: {} for _, header in réglage_cols}
    conforme_summary = {header: 0 for _, header in amortissement_cols}
    controlled_summary = {header: 0 for _, header in amortissement_cols}  # Compteur pour les FKT contrôlées
    reglage_conforme = {header: {} for _, header in amortissement_cols}
    total_count = {header: 0 for _, header in réglage_cols}
    conforme_par_reglage = {header: 0 for _, header in réglage_cols}  # Ajout du compteur par colonne de réglage
    data = []
    total_rows = 0
    actual_fkt_count = 0  # Nouveau compteur pour les FKT réelles
    
    # Pour les tendances - suivi par semaine ou mois
    fkt_timeline = {}
    
    # Dictionnaire pour stocker les détails des conformités
    conformite_details = {header: {"details": [], "values": []} for _, header in amortissement_cols}
    
    # Traitement des données
    last_valid_row = 0  # Pour suivre la dernière ligne contenant des données valides
    
    main_colors = []  # 🆕 pour stocker la couleur principale de chaque ligne FKT

    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row+1, values_only=False)):
        # Vérifier si la ligne est vide ou ne contient que des cellules sans valeur
        if all((cell is None or str(cell.value).strip() == "") for cell in row):
            continue

            
        # Vérifier si la ligne contient des données FKT réelles
        has_fkt_data = False
        for col_idx, _ in réglage_cols:
            if col_idx < len(row) and row[col_idx].value is not None:
                has_fkt_data = True
                last_valid_row = row_idx
                break
                
        if has_fkt_data:
            actual_fkt_count += 1
        
        total_rows += 1
        row_data = {}
        
        # Extraire les données et couleurs en une seule passe
        row_values = []
        row_colors = []
        
        for cell in row:
            row_values.append(cell.value)

            try:
                # Meilleure compatibilité : on tente d'abord fgColor
                fill_color = cell.fill.fgColor if cell.fill.fgColor and cell.fill.fgColor.rgb else cell.fill.start_color
                color_code = safe_extract_hex(fill_color)


            except:
                color_code = 'FFFFFF'

            row_colors.append(get_standardized_color(color_code))

        # Construire un dictionnaire pour la ligne
        for i, header in enumerate(headers):
            if i < len(row_values):
                row_data[header] = row_values[i]
        
        # Ajouter aux données
        data.append(row_data)
        
        # Traiter les colonnes de réglage
        row_color_candidates = []

        for col_idx, header in réglage_cols:
            if col_idx < len(row_values):
                value = row_values[col_idx]

                if value is None:
                    continue

                total_count[header] += 1
                raw_color = row_colors[col_idx]
                standardized_color = get_standardized_color(raw_color)

                color_summary[header][standardized_color] = color_summary[header].get(standardized_color, 0) + 1

                # 🆕 on ajoute cette couleur à la liste des couleurs candidates de la ligne
                row_color_candidates.append(standardized_color)

        # Choisir une seule couleur dominante pour cette ligne
        main_color = 'FFFFFF'  # par défaut
        for c in row_color_candidates:
            if c != 'FFFFFF':
                main_color = c
                break
        main_colors.append(main_color)
        
        # Traiter les colonnes d'amortissement
        conforme_found = False  # Pour suivre si une valeur conforme a été trouvée dans cette ligne
        conforme_reglage_col = None  # Pour suivre quelle colonne de réglage est associée à la conformité
        
        for col_idx, header in amortissement_cols:
            if col_idx < len(row_values):
                value = row_values[col_idx]
                
                # Stocker les valeurs d'amortissement pour analyse
                if value is not None:
                    conformite_details[header]["values"].append(str(value))
                
                # Vérifier si des FKT ont été contrôlées pour cette ligne
                fkt_controlled = False
                for reglage_col, reglage_header in réglage_cols:
                    if reglage_col < len(row_colors):
                        reglage_color = get_standardized_color(row_colors[reglage_col])
                        # 00B0F0 = FKT contrôlée sur Plan, 92CDDC = FKT contrôlée sur Terrain
                        if reglage_color in ['00B0F0', '92CDDC', '92D050']:
                            fkt_controlled = True
                            conforme_reglage_col = (reglage_col, reglage_header)  # Stocker la colonne de réglage pour cette conformité
                            break
                
                # Comptage seulement pour les FKT contrôlées
                # Nouvelle version : on compte toutes les lignes renseignées, même sans contrôle
                conformite = is_conforme(value)

                if conformite is not None:
                    controlled_summary[header] += 1  # Compter comme FKT renseignée

                    if conformite is True:
                        conforme_summary[header] += 1

                    conformite_details[header]["details"].append({"value": str(value), "conforme": conformite})
                else:
                    conformite_details[header]["details"].append({"value": str(value), "conforme": None})


                    # Comptage par colonne de réglage
                    for reglage_col, reglage_header in réglage_cols:
                        if reglage_col < len(row_values):
                            reglage_value = row_values[reglage_col]
                            reglage_color = row_colors[reglage_col]
                            
                            # Standardiser la couleur
                            reglage_color = get_standardized_color(reglage_color)
                            
                            if reglage_value == 1 and reglage_color == 'FFFFFF':
                                if reglage_header not in reglage_conforme[header]:
                                    reglage_conforme[header][reglage_header] = 0
                                reglage_conforme[header][reglage_header] += 1
        
        # Si une valeur conforme a été trouvée dans cette ligne, incrémenter le compteur pour la colonne de réglage associée
        if conforme_found and conforme_reglage_col is not None:
            _, reglage_header = conforme_reglage_col
            conforme_par_reglage[reglage_header] += 1
    
    # Conversion en DataFrame
    df = pd.DataFrame(data)
    
    # Conversion de fkt_timeline en format adapté pour Chart.js
    sorted_weeks = sorted(fkt_timeline.keys())
    timeline_data = {
        "labels": [f"Semaine {week}" for week in sorted_weeks],
        "datasets": []
    }
    
    # Ajout d'un jeu de données pour chaque couleur FKT
    for color_code, info in color_map.items():
        dataset = {
            "label": info['desc'],
            "backgroundColor": info['hex'],
            "borderColor": info['hex'],
            "data": [fkt_timeline.get(week, {}).get(color_code, 0) for week in sorted_weeks]
        }
        timeline_data["datasets"].append(dataset)
    
    # Créer des données pour la progression
    progression_data = {
        "labels": ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"],
        "datasets": [
            {
                "label": "Progression des contrôles",
                "data": [
                    random.randint(20, 50),
                    random.randint(30, 60),
                    random.randint(40, 70),
                    random.randint(50, 80),
                    random.randint(60, 90),
                    random.randint(70, 100),
                    random.randint(60, 90),
                    random.randint(70, 100),
                    random.randint(80, 110),
                    random.randint(90, 120),
                    random.randint(100, 130),
                    random.randint(110, 140)
                ],
                "fill": False,
                "borderColor": "rgb(75, 192, 192)",
                "tension": 0.1
            }
        ]
    }
    
    doughnut_data = {
        "labels": [color_map[color]['desc'] for color in color_map],
        "datasets": [
            {
                "label": "FKT par statut",
                "data": [sum(color_summary[header].get(color, 0) for header in [h for _, h in réglage_cols]) for color in color_map],
                "backgroundColor": [color_map[color]['hex'] for color in color_map],
                "hoverOffset": 4
            }
        ]
    }

    
    # Utiliser actual_fkt_count au lieu de total_rows pour les FKT réelles
    return (df, color_summary, conforme_summary, controlled_summary, réglage_cols, amortissement_cols, 
        actual_fkt_count, total_count, reglage_conforme, timeline_data, progression_data, 
        doughnut_data, conformite_details, conforme_par_reglage)


def create_pie_chart(data, title, total_count=None):
    """Crée un graphique en camembert avec couleurs & labels 100% fidèles au color_map"""

    # Étape 1 : regrouper les couleurs standardisées
    color_counts = {}
    for raw_color, count in data.items():
        standardized = get_standardized_color(raw_color)
        color_counts[standardized] = color_counts.get(standardized, 0) + count

    # Étape 2 : créer les données triées selon l'ordre dans color_map
    labels = []
    values = []
    colors = []

    for code, meta in color_map.items():
        labels.append(meta['desc'])
        values.append(color_counts.get(code, 0))
        colors.append(meta['hex'])

    # ✅ Étape 3 : filtrer les graphiques vides ou 100% à venir
    total = sum(values)
    index_fkt_avenir = labels.index("FKT à venir")
    fkt_avenir_only = (total > 0 and values[index_fkt_avenir] == total)

    if total == 0 or fkt_avenir_only:
        return None

    # Étape 4 : créer le graphique avec les couleurs manuellement assignées
    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        marker=dict(colors=colors),
        textinfo='percent+label',
        insidetextorientation='radial'
    )])

    fig.update_layout(
    title=dict(
        text=title,
        x=0.5,
        xanchor="center",
        y=0.95,       # ⬅⬅⬅ Position plus basse que le haut
        yanchor="top"
    ),
    legend=dict(
        orientation="h",
        yanchor="bottom",
        y=-0.3,
        xanchor="center",
        x=0.5
    ),
    margin=dict(t=120, b=100, l=20, r=20),  # t=120 pour faire descendre le titre
    height=550
)


    if total_count:
        fig.add_annotation(
            text=f"Total: {total_count}",
            x=0.5, y=1.1,
            xref="paper", yref="paper",
            showarrow=False,
            font=dict(size=14)
        )

    return fig

def create_bar_chart(data, title, total_count=None):
    """Crée un graphique à barres optimisé"""
    labels = []
    values = []
    colors = []
    
    for color_code in color_map:
        labels.append(color_map[color_code]['desc'])
        colors.append(color_map[color_code]['hex'])
        values.append(data.get(color_code, 0))

    
    if sum(values) == 0:
        fig = go.Figure()
        fig.update_layout(
            title=f"{title} - Aucune donnée",
            annotations=[dict(text="Aucune donnée", showarrow=False, font=dict(size=20))]
        )
        return fig
    
    fig = px.bar(
        x=labels, 
        y=values, 
        title=title,
        color=labels,
        color_discrete_map={label: color for label, color in zip(labels, colors)}
    )
    
    fig.update_layout(
        xaxis_title="Statut",
        yaxis_title="Nombre",
        showlegend=False,
        height=350
    )
    
    if total_count:
        fig.add_annotation(
            text=f"Total: {total_count}",
            x=0.5, y=1.05,
            xref="paper", yref="paper",
            showarrow=False,
            font=dict(size=14)
        )
    
    return fig

def create_radar_chart(data):
    """Crée un graphique radar pour visualiser la distribution des FKT par statut"""
    
    # Extraire les données du dictionnaire global doughnut_data
    labels = data["labels"]
    values = data["datasets"][0]["data"]
    colors = data["datasets"][0]["backgroundColor"]
    
    # Créer le graphique radar
    fig = go.Figure()
    
    fig.add_trace(go.Scatterpolar(
        r=values,
        theta=labels,
        fill='toself',
        name='Statut FKT',
        line=dict(color='rgba(32, 128, 128, 0.8)'),
        fillcolor='rgba(32, 128, 128, 0.3)'
    ))
    
    fig.update_layout(
        polar=dict(
            radialaxis=dict(
                visible=True,
                range=[0, max(values) * 1.1 if values else 1]
            )
        ),
        title="Distribution des FKT par statut",
        height=450
    )
    
    return fig

def create_fkt_controlled_pie_chart(total_fkt, total_controlled):
    """Affiche le ratio FKT contrôlées / non contrôlées"""
    non_controlled = total_fkt - total_controlled
    
    fig = go.Figure(data=[go.Pie(
        labels=["FKT contrôlées", "FKT non contrôlées"],
        values=[total_controlled, non_controlled],
        marker=dict(colors=["#00B0F0", "#D3D3D3"]),
        hole=0.4
    )])
    
    fig.update_layout(
        title="Répartition des FKT contrôlées",
        annotations=[dict(text=f"{total_controlled}/{total_fkt}", x=0.5, y=0.5, showarrow=False, font_size=18)],
        height=400,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
    )
    
    return fig

def create_conforme_gauge(value, total, title):
    """Crée un indicateur de type jauge optimisé"""
    percentage = (value / total) * 100 if total > 0 else 0
    
    fig = go.Figure(go.Indicator(
        mode="gauge+number",
        value=percentage,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': title},
        number={'suffix': '%'},
        gauge={
            'axis': {'range': [0, 100]},
            'bar': {'color': "#00B050"},
            'steps': [
                {'range': [0, 33], 'color': "#FF0000"},
                {'range': [33, 66], 'color': "#FFFF00"},
                {'range': [66, 100], 'color': "#92D050"}
            ],
            'threshold': {
                'line': {'color': "black", 'width': 4},
                'thickness': 0.75,
                'value': 80
            }
        }
    ))
    
    fig.update_layout(
        height=250,
        margin=dict(t=50, b=30, l=30, r=30)
    )
    
    fig.add_annotation(
        text=f"{value} / {total}",
        x=0.5, y=0.3,
        xref="paper", yref="paper",
        showarrow=False,
        font=dict(size=14)
    )
    
    return fig

def create_conformite_details_chart(details_data):
    """Crée un graphique pour visualiser les détails des valeurs de conformité"""
    if not details_data or "values" not in details_data or not details_data["values"]:
        return None
    
    # Analyser les valeurs pour les catégoriser
    categories = {}
    for value in details_data["values"]:
        value = str(value).strip().lower()
        # Simplifier et regrouper les valeurs similaires
        if any(term in value for term in ["conforme", "ok", "validé", "oui"]):
            key = "Conforme"
        elif any(term in value for term in ["non conforme", "non-conforme", "écart", "ko", "non"]):
            key = "Non Conforme"
        elif value == "":
            key = "Non renseigné"
        else:
            key = "Autre"
        
        categories[key] = categories.get(key, 0) + 1
    
    # Créer le graphique avec les catégories
    fig = px.pie(
        names=list(categories.keys()),
        values=list(categories.values()),
        title="Catégorisation des valeurs de conformité",
        color_discrete_sequence=px.colors.qualitative.Set3
    )
    
    fig.update_traces(textposition='inside', textinfo='percent+label')
    fig.update_layout(height=350)
    
    return fig

def create_amortissement_conformity_charts(controlled_summary, conforme_summary, conformite_details=None):
    """Crée des graphiques pour visualiser le taux de conformité d'amortissement"""

    # Extraire toutes les valeurs depuis conformite_details
    conforme_vals = []
    for details in (conformite_details or {}).values():
        conforme_vals.extend([item["conforme"] for item in details.get("details", [])])

    # Compter les types
    total_conforme = sum(1 for v in conforme_vals if v is True)
    total_non_conforme = sum(1 for v in conforme_vals if v is False)
    total_non_renseigne = sum(1 for v in conforme_vals if v is None)
    total_controlled = total_conforme + total_non_conforme  # ignore les "non renseigné"

    # Calcul du taux
    if total_controlled > 0:
        conformity_rate = (total_conforme / total_controlled) * 100
    else:
        conformity_rate = 0


    # Graphique en anneau avec les 3 catégories
    fig_pie = go.Figure(data=[go.Pie(
        labels=["Conforme", "Non Conforme", "Non renseigné"],
        values=[
            total_conforme,
            total_non_conforme,
            total_controlled - total_conforme - total_non_conforme
        ],
        hole=.4,
        marker_colors=["#00B050", "#FF0000", "#D3D3D3"]
    )])
    fig_pie.update_layout(
        title=f"Taux Global de Conformité: {conformity_rate:.1f}%",
        annotations=[dict(
            text=f"{total_conforme}/{total_controlled}",
            x=0.5, y=0.5, font_size=20, showarrow=False
        )],
        height=400
    )

    # Graphique à barres horizontales par type
    fig_horizontal = go.Figure()

    labels = list(conforme_summary.keys())
    conforme_values = list(conforme_summary.values())
    controlled_values = list(controlled_summary.values())

    percentages = [
        (conf / ctrl) * 100 if ctrl > 0 else 0 
        for conf, ctrl in zip(conforme_values, controlled_values)
    ]

    fig_horizontal.add_trace(go.Bar(
        y=labels,
        x=percentages,
        orientation='h',
        marker_color='rgba(0, 176, 80, 0.7)',
        name='Taux de Conformité'
    ))

    fig_horizontal.update_layout(
        title="Taux de Conformité par Type",
        xaxis_title="Pourcentage (%)",
        yaxis_title="Type d'Amortissement",
        height=400,
        xaxis=dict(range=[0, 100])
    )

    # Graphique temporel mensuel (simulé)
    fig_monthly = go.Figure()

    months = ["Jan", "Fév", "Mar", "Avr", "Mai", "Jun", "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]
    random_values = [random.uniform(60, 95) for _ in range(12)]

    fig_monthly.add_trace(go.Scatter(
        x=months,
        y=random_values,
        mode='lines+markers',
        name='Évolution Mensuelle',
        line=dict(color='rgb(0, 176, 240)', width=3)
    ))

    fig_monthly.update_layout(
        title="Évolution Mensuelle du Taux de Conformité",
        xaxis_title="Mois",
        yaxis_title="Taux de Conformité (%)",
        height=400,
        yaxis=dict(range=[50, 100])
    )

    # Graphique cumulatif (simulé)
    fig_cumulative = go.Figure()

    cumulative_data = [sum(random_values[:i+1])/(i+1) for i in range(12)]

    fig_cumulative.add_trace(go.Scatter(
        x=months,
        y=cumulative_data,
        mode='lines+markers',
        fill='tozeroy',
        name='Cumulatif Annuel',
        line=dict(color='rgb(146, 208, 80)', width=3),
        fillcolor='rgba(146, 208, 80, 0.3)'
    ))

    fig_cumulative.update_layout(
        title="Évolution Cumulative du Taux de Conformité",
        xaxis_title="Mois",
        yaxis_title="Taux de Conformité Cumulé (%)",
        height=400,
        yaxis=dict(range=[50, 100])
    )

    return fig_monthly, fig_cumulative, fig_pie, fig_horizontal

def create_progress_chart(data):
    """Crée un graphique linéaire pour montrer la progression des contrôles"""
    fig = go.Figure()
    
    for dataset in data["datasets"]:
        fig.add_trace(go.Scatter(
            x=data["labels"],
            y=dataset["data"],
            mode='lines+markers',
            name=dataset["label"],
            line=dict(color=dataset["borderColor"], width=3),
            marker=dict(size=8)
        ))
    
    fig.update_layout(
        title="Progression des contrôles dans le temps",
        xaxis_title="Période",
        yaxis_title="Nombre de contrôles",
        height=400
    )
    
    return fig

def create_doughnut_chart(data, title="Répartition globale des FKT"):
    """Crée un graphique en anneau pour la répartition des FKT"""
    fig = go.Figure(data=[go.Pie(
        labels=data["labels"],
        values=data["datasets"][0]["data"],
        hole=.4,
        marker_colors=data["datasets"][0]["backgroundColor"]
    )])
    
    total = sum(data["datasets"][0]["data"])
    
    fig.update_layout(
        title=title,
        annotations=[dict(text=f"Total: {total}", x=0.5, y=0.5, font_size=20, showarrow=False)],height=400,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
    )
    
    return fig

def create_timeline_chart(data):
    """Crée un graphique de tendance pour l'évolution des FKT dans le temps"""
    fig = go.Figure()
    
    for dataset in data["datasets"]:
        fig.add_trace(go.Bar(
            x=data["labels"],
            y=dataset["data"],
            name=dataset["label"],
            marker_color=dataset["backgroundColor"]
        ))
    
    fig.update_layout(
        title="Évolution des FKT dans le temps",
        xaxis_title="Période",
        yaxis_title="Nombre de FKT",
        barmode='stack',
        height=450,
        legend=dict(orientation="h", yanchor="bottom", y=-0.3, xanchor="center", x=0.5)
    )
    
    return fig

def get_progress_metrics(color_summary, actual_fkt_count=None):
    # Utiliser uniquement le nombre de FKT réelles détectées (une par ligne)
    total_fkt = actual_fkt_count if actual_fkt_count is not None else 0

    # FKT Réalisées = toutes sauf rouge et blanc
    total_realized = sum(
        sum(count for color, count in counts.items() if color not in ['FF0000', 'FFFFFF'])
        for counts in color_summary.values()
    )

    # FKT contrôlées = plan + terrain + réalisées
    total_controlled = sum(
        counts.get('00B0F0', 0) + counts.get('92CDDC', 0) + counts.get('92D050', 0)
        for counts in color_summary.values()
    )

    total_with_deviations = sum(
        counts.get('FF0000', 0) for counts in color_summary.values()
    )

    # Ici, on ne divise pas par la somme des couleurs (fausse) mais bien par actual_fkt_count
    if total_fkt == 0:
        percent_realized = 0
        percent_controlled = 0
        percent_with_deviations = 0
    else:
        percent_realized = (total_realized / total_fkt) * 100
        percent_controlled = (total_controlled / total_fkt) * 100
        percent_with_deviations = (total_with_deviations / total_controlled) * 100 if total_controlled > 0 else 0

    return {
        "total_fkt": total_fkt,
        "total_realized": total_realized,
        "total_controlled": total_controlled,
        "total_with_deviations": total_with_deviations,
        "percent_realized": percent_realized,
        "percent_controlled": percent_controlled,
        "percent_with_deviations": percent_with_deviations
    }


def generate_json_for_export(color_summary, conforme_summary, controlled_summary, 
                            total_rows, doughnut_data, timeline_data):
    """Génère un JSON avec toutes les données importantes pour l'export"""
    
    # Obtenir les métriques de progression
    progress_metrics = get_progress_metrics(color_summary, actual_fkt_count=total_rows)


    
    # Préparer les données par catégorie de FKT
    fkt_by_status = {}
    for color_code, info in color_map.items():
        fkt_by_status[info['desc']] = sum(
            counts.get(color_code, 0) for counts in color_summary.values()
        )
    
    # Calcul du taux de conformité global
    total_conforme = sum(conforme_summary.values())
    total_controlled_amort = sum(controlled_summary.values())
    conformity_rate = (total_conforme / total_controlled_amort) * 100 if total_controlled_amort > 0 else 0
    
    # Créer l'objet JSON final
    export_data = {
        "date_generation": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": {
            "total_rows": total_rows,
            "total_fkt": progress_metrics["total_fkt"],
            "fkt_by_status": fkt_by_status,
            "conformity": {
                "rate": round(conformity_rate, 2),
                "conforme": total_conforme,
                "controlled": total_controlled_amort
            },
            "progress": {
                "percent_realized": round(progress_metrics["percent_realized"], 2),
                "percent_controlled": round(progress_metrics["percent_controlled"], 2),
                "percent_with_deviations": round(progress_metrics["percent_with_deviations"], 2)
            }
        },
        "charts_data": {
            "doughnut": doughnut_data,
            "timeline": timeline_data
        }
    }
    
    return export_data

def generate_dashboard_pdf(progress_metrics, doughnut_data, fig_conformity=None, df=None, detail_figs=None, control_fig=None):
    """Génère un PDF complet avec toutes les visualisations de l'application FKT."""

    # Date actuelle pour l'en-tête
    now = datetime.now().strftime("%d/%m/%Y")

    # Partie 1 : Structure HTML avec entête UTF-8
    html_parts = [
        f"<h1 style='text-align: center;'>Rapport de Suivi des FKT</h1>",
        f"<p style='text-align: right;'><em>Date de génération : {now}</em></p>",
        "<hr>",
        "<h2>Résumé des Indicateurs Clés</h2>",
        "<ul>",
        f"<li><strong>Nombre total de FKT :</strong> {progress_metrics['total_fkt']}</li>",
        f"<li><strong>FKT réalisées :</strong> {progress_metrics['total_realized']} ({progress_metrics['percent_realized']:.1f}%)</li>",
        f"<li><strong>FKT contrôlées :</strong> {progress_metrics['total_controlled']} ({progress_metrics['percent_controlled']:.1f}%)</li>",
        f"<li><strong>FKT avec écarts :</strong> {progress_metrics['total_with_deviations']} ({progress_metrics['percent_with_deviations']:.1f}%)</li>",
        "</ul>",
        "<hr>"
    ]

    # Fonction pour transformer un graphique Plotly en base64
    def fig_to_base64(fig):
        img_bytes = fig.to_image(format="png", width=800, height=500, engine="kaleido")
        base64_img = base64.b64encode(img_bytes).decode("utf-8")
        return f'<img src="data:image/png;base64,{base64_img}" width="700"><br><br>'

    # SECTION : Vue d'ensemble
    html_parts.append("<h2>1. Vue d’ensemble</h2>")
    html_parts.append("<h3>1.1 Répartition globale des FKT</h3>")
    html_parts.append(fig_to_base64(pio.from_json(pio.to_json(create_doughnut_chart(doughnut_data)))))

    if control_fig:
        html_parts.append("<h3>1.2 Répartition des FKT contrôlées</h3>")
        html_parts.append(fig_to_base64(pio.from_json(pio.to_json(control_fig))))
    
    # SECTION : Conformité
    if fig_conformity:
        html_parts.append("<h2>2. Analyse de la Conformité</h2>")
        html_parts.append("<h3>2.1 Taux global de conformité</h3>")
        html_parts.append(fig_to_base64(pio.from_json(pio.to_json(fig_conformity))))

    # SECTION : Détails par type de FKT
    if detail_figs:
        html_parts.append("<h2>3.</h2>")
        for label, fig in detail_figs:
            if fig is None:
                continue  # ⛔ ne pas inclure les figures vides
            html_parts.append(f"<h4>{label}</h4>")
            html_parts.append(fig_to_base64(pio.from_json(pio.to_json(fig))))


    # HTML final avec encodage
    full_html = (
        "<html><head><meta charset='UTF-8'></head>"
        "<body style='font-family: Arial, sans-serif;'>"
        + "".join(html_parts) +
        "</body></html>"
    )

    # Partie 2 : Génération du PDF avec wkhtmltopdf
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmpfile:
        pdfkit_config = None
        pdfkit.from_string(full_html, tmpfile.name) if pdfkit_config is None else pdfkit.from_string(full_html, tmpfile.name, configuration=pdfkit_config)
        return tmpfile.name



def main():
    """Fonction principale de l'application Streamlit"""
    
    st.set_page_config(
        page_title="Dashboard FKT",
        page_icon="",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    st.title("Tableau de Bord de Suivi des FKT")

    with st.sidebar:
        st.header("Configuration")
        uploaded_file = st.file_uploader("Choisir un fichier Excel", type=["xlsx", "xls"])

        with st.expander("Options avancées", expanded=False):
            auto_detect_headers = st.checkbox("Détection automatique des en-têtes", value=True)
            fallback_row = st.number_input("Ligne de départ (si détection auto désactivée)", min_value=1, value=1)
        
        st.markdown("---")
        st.markdown("### Légende")
        for code, info in color_map.items():
            st.markdown(
                f"<div style='display: flex; align-items: center;'>"
                f"<div style='width: 20px; height: 20px; background-color: {info['hex']}; "
                f"border: 1px solid #ddd; margin-right: 10px;'></div>"
                f"<div>{info['desc']}</div>"
                f"</div>",
                unsafe_allow_html=True
            )

        st.markdown("---")

    if uploaded_file is not None:
        try:
            file_bytes = uploaded_file.read()
            sheet_names = get_sheet_names(file_bytes)
            selected_sheet = st.selectbox("Sélectionner une feuille", sheet_names)

            # Étape 1 : prévisualisation des lignes
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = wb[selected_sheet] if selected_sheet in wb.sheetnames else wb.active

            st.subheader("Aperçu des premières lignes du fichier Excel")
            preview_data = []
            max_preview_rows = 20
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_preview_rows, values_only=True), start=1):
                preview_data.append([i] + list(row))
            df_preview = pd.DataFrame(preview_data)
            df_preview.columns = ["Ligne"] + [f"Col{i}" for i in range(1, len(df_preview.columns))]
            st.dataframe(df_preview)

            # Étape 2 : choix de la ligne d'en-tête
            selected_header_row = st.selectbox("Sélectionne la ligne à utiliser comme en-tête :", options=df_preview["Ligne"])
            start_row = selected_header_row
            auto_detect_headers = False

            # Étape 3 : chargement des données
            result = load_targeted_data(file_bytes, selected_sheet, start_row, auto_detect_headers)
            (df, color_summary, conforme_summary, controlled_summary, réglage_cols, amortissement_cols,
             total_rows, total_count, reglage_conforme, timeline_data, progression_data,
             doughnut_data, conformite_details, conforme_par_reglage) = result

            # Étape 4 : sélection manuelle des colonnes à comparer
            st.subheader("Sélection manuelle des colonnes à analyser")
            selected_reglage_col = st.selectbox("Choisir la colonne de Réglage :", options=df.columns)
            selected_amortissement_cols = st.multiselect(
                " Choisis une ou plusieurs colonnes d'Amortissement à analyser :",
                options=df.columns,
                default=[col for col in df.columns if "amortissement" in col.lower()]
            )


            st.success(f"Tu as choisi : Réglage = **{selected_reglage_col}**, Amortissement = **{', '.join(selected_amortissement_cols)}**")

            # ➕ ici tu peux utiliser selected_reglage_col et selected_amortissement_col dans un traitement dédié

            export_data = generate_json_for_export(
                color_summary, conforme_summary, controlled_summary, 
                total_rows, doughnut_data, timeline_data
            )
            st.session_state.export_data = export_data

            progress_metrics = get_progress_metrics(color_summary, actual_fkt_count=total_rows)

            tab1, tab2, tab3, tab4 = st.tabs(["Vue d'ensemble", "Détails FKT", "Conformité", "Données brutes"])

            with tab1:
                # 🔹 1. Créer le graphe des FKT contrôlées
                fig_controlled = create_fkt_controlled_pie_chart(
                    progress_metrics["total_fkt"],
                    progress_metrics["total_controlled"]
                )

                # 🔹 2. Créer le graphe de conformité global (camembert)
                _, _, fig_conformity, _ = create_amortissement_conformity_charts(
                    controlled_summary,
                    conforme_summary,
                    conformite_details
                )

                # 🔹 3. Créer les graphiques par type de FKT
                detail_figs = [
                    (
                        f"Répartition - {header}",
                        create_pie_chart(color_summary[header], f"Répartition - {header}", total_count[header])
                    )
                    for _, header in réglage_cols
                ]
                detail_figs = [(label, fig) for label, fig in detail_figs if fig is not None]

                # 🔹 4. Générer le PDF avec tous les éléments
                pdf_path = generate_dashboard_pdf(
                    progress_metrics=progress_metrics,
                    doughnut_data=doughnut_data,
                    fig_conformity=fig_conformity,
                    df=df,
                    detail_figs=detail_figs,
                    control_fig=fig_controlled
                )

                # 🔹 5. Télécharger le fichier PDF
                with open(pdf_path, "rb") as f:
                    st.download_button(
                        label="📥 Télécharger le PDF complet",
                        data=f,
                        file_name="rapport_FKT.pdf",
                        mime="application/pdf"
                    )


                col1, col2, col3, col4 = st.columns(4)
                col1.metric("Total FKT", progress_metrics["total_fkt"])
                col2.metric("FKT Réalisées", f"{progress_metrics['total_realized']} ({progress_metrics['percent_realized']:.1f}%)")
                col3.metric("FKT Contrôlées", f"{progress_metrics['total_controlled']} ({progress_metrics['percent_controlled']:.1f}%)")
                col4.metric("FKT avec écarts", f"{progress_metrics['total_with_deviations']} ({progress_metrics['percent_with_deviations']:.1f}%)")

                col1, col2 = st.columns(2)
                st.plotly_chart(create_doughnut_chart(doughnut_data, "Répartition globale des FKT"), use_container_width=True)
                st.plotly_chart(create_fkt_controlled_pie_chart(progress_metrics["total_fkt"], progress_metrics["total_controlled"]), use_container_width=True)

            with tab2:
                st.subheader("Détails par type de FKT")
                
                for i in range(0, len(detail_figs), 2):
                    cols = st.columns([0.45, 0.45, 0.10])

                    for j in range(2):
                        if i + j < len(detail_figs):
                            _, fig = detail_figs[i + j]
                            with cols[j]:
                                st.plotly_chart(fig, use_container_width=True)

                    # ➕ Espace vertical entre chaque paire de graphiques
                    st.markdown("<br><br>", unsafe_allow_html=True)

                        
            with tab3:
                st.subheader("Analyse de la Conformité")
                total_conforme = sum(conforme_summary.values())
                total_controlled = sum(controlled_summary.values())

                if total_controlled > 0:
                    _, _, fig_pie, _ = create_amortissement_conformity_charts(controlled_summary, conforme_summary, conformite_details)
                    st.plotly_chart(fig_pie, use_container_width=True)
                else:
                    st.info("Aucune FKT contrôlée pour calculer la conformité")

            with tab4:
                st.subheader("Données brutes")
                st.write("Aperçu des données :")
                st.dataframe(df.head(20))

                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Télécharger les données CSV",
                    data=csv,
                    file_name="fkt_data.csv",
                    mime="text/csv"
                )

        except Exception as e:
            st.error(f"Erreur lors du traitement du fichier : {str(e)}")
            st.exception(e)
    else:
        st.info("Veuillez charger un fichier Excel contenant vos données FKT dans le panneau de gauche.")

        
        # Afficher une présentation quand aucun fichier n'est chargé
        col1, col2 = st.columns([1, 1])
        
        with col1:
            st.markdown("""
            ### Fonctionnalités principales
            - Visualisation immédiate de la répartition des FKT
            - Analyse détaillée par type de FKT
            - Suivi des taux de conformité
            - Analyse des écarts
            - Interface responsive
            - Export des données et résultats
            
            ### Comment utiliser ce tableau de bord
            1. Chargez votre fichier Excel depuis le panneau de gauche
            2. Sélectionnez la feuille contenant vos données
            3. Explorez les différents onglets pour analyser vos données
            4. Téléchargez les résultats au format CSV ou JSON
            """)
        
        with col2:
            # Simuler un exemple de graphique
            data = {
                "labels": ["FKT à venir", "FKT réalisée", "FKT contrôlée sur Plan", 
                           "FKT contrôlée sur Terrain", "FKT avec écart"],
                "datasets": [{
                    "label": "Exemple",
                    "data": [45, 30, 15, 8, 2],
                    "backgroundColor": ["#FFFFFF", "#92D050", "#00B0F0", "#92CDDC", "#FF0000"]
                }]
            }
            
            # Créer et afficher un graphique d'exemple
            example_fig = create_doughnut_chart(data, "Exemple de visualisation")
            st.plotly_chart(example_fig, use_container_width=True)

if __name__ == "__main__":
    main()
